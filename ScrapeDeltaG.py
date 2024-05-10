from httpcore import TimeoutException
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

def get_driver():
    chrome_path = '/Users/noahnicol/Desktop/Desktop/Job_Stuff/Proteovista/chromedriver-mac-arm64/chromedriver'
    service = Service(executable_path=chrome_path)
    chrome_options = Options()
    chrome_options.add_argument("--start-fullscreen")  # This sets the browser to start in full screen
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

# Use this function to initialize your driver
driver = get_driver()


def read_sequences_from_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sequences = []
    labels = []
    
    # Assuming the first row contains data and there is a header
    row = 2
    while True:
        label = sheet[f'E{row}'].value
        sequence = sheet[f'F{row}'].value
        # Break the loop if there's no more data
        if label is None or sequence is None:
            break
        labels.append(label)
        sequences.append(sequence)
        row += 1

    workbook.close()
    print("sequences:", sequences)
    print("labels:", labels)
    return sequences, labels

counter = 0
isLoggedIn = False

def analyze_dimer(driver, sequence1, sequence2, **kwargs):
    global isLoggedIn  # Ensure you're using the global variable
    try:
        if not isLoggedIn:
            login_to_idt(driver)
            isLoggedIn = True  # Update the login status after a successful login

        # Navigate to the Oligo Analyzer page
        driver.get("https://www.idtdna.com/calc/analyzer")

        # Check if the main sequence area is accessible, on error we're redirected to login
        main_sequence_area = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "textarea-sequence"))
        )
        main_sequence_area.clear()
        main_sequence_area.send_keys(sequence1)

        # Set the concentrations if specified in kwargs
        oligo = kwargs.get('oligo', 'default')
        na = kwargs.get('na', 'default')
        mg = kwargs.get('mg', 'default')
        dntp = kwargs.get('dntp', 'default')

        if 'oligo' != 'default':
            oligo_input = driver.find_element(By.CSS_SELECTOR, "input[data-bind*='oligoConc']")
            oligo_input.clear()
            oligo_input.send_keys(str(oligo))

        if 'na' != 'default':
            na_input = driver.find_element(By.CSS_SELECTOR, "input[data-bind*='naConc']")
            na_input.clear()
            na_input.send_keys(str(na))

        if 'mg' != 'default':
            mg_input = driver.find_element(By.CSS_SELECTOR, "input[data-bind*='mgConc']")
            mg_input.clear()
            mg_input.send_keys(str(mg))

        if 'dntp' != 'default':
            dntp_input = driver.find_element(By.CSS_SELECTOR, "input[data-bind*='dNTPsConc']")
            dntp_input.clear()
            dntp_input.send_keys(str(dntp))

        hetero_dimer_button = WebDriverWait(driver, 12).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='rmenu']//button[contains(text(), 'Hetero-Dimer')]"))
        )
        hetero_dimer_button.click()

        secondary_sequence_box = WebDriverWait(driver, 25).until(
            EC.presence_of_element_located((By.XPATH, "//textarea[contains(@data-bind, 'secondarySequence')]"))
        )
        # print("secondary sequence box found")
        secondary_sequence_box.clear()
        secondary_sequence_box.send_keys(sequence2)

        calculate_button = WebDriverWait(driver, 18).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@data-bind, 'click: calculate1')]"))
        )
        calculate_button.click()

        delta_g_value = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, "//span[contains(@data-bind, 'text: deltaG')]"))
        )
        delta_g = delta_g_value.text.split()[0]  # Extracts the delta G value from format "-X.XX kcal/mol"
        print("Delta G:", float(delta_g))
        global counter
        counter += 1
        return float(delta_g)
    except Exception as e:
        print("An error occurred in analyze_dimer:", str(e))
        print("count:", counter)
        isLoggedIn = False  # Reset login status on failure
        raise
    

def analyze_hairpin(driver, sequence, **kwargs):
    try:
        # Navigate to the Oligo Analyzer page
        driver.get("https://www.idtdna.com/calc/analyzer")
        
        # Input the sequence
        sequence_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "textarea-sequence"))
        )
        sequence_input.clear()
        sequence_input.send_keys(sequence)
        
        # Set the concentrations if specified in kwargs and not 'default'
        oligo = kwargs.get('oligo', 'default')
        na = kwargs.get('na', 'default')
        mg = kwargs.get('mg', 'default')
        dntp = kwargs.get('dntp', 'default')

        if oligo != 'default':
            oligo_input = driver.find_element(By.CSS_SELECTOR, "input[data-bind*='oligoConc']")
            oligo_input.clear()
            oligo_input.send_keys(str(oligo))

        if na != 'default':
            na_input = driver.find_element(By.CSS_SELECTOR, "input[data-bind*='naConc']")
            na_input.clear()
            na_input.send_keys(str(na))

        if mg != 'default':
            mg_input = driver.find_element(By.CSS_SELECTOR, "input[data-bind*='mgConc']")
            mg_input.clear()
            mg_input.send_keys(str(mg))

        # Click the Hairpin button
        hairpin_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "hairpin-button"))
        )
        hairpin_button.click()

        # Wait for the Delta G value to be visible
        delta_g_value = WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "td[data-bind='text: deltaG']"))
        )
        delta_g = delta_g_value.text.strip()  # Extracts the delta G value
        
        print("Delta G for Hairpin:", delta_g)
        return float(delta_g)

    except Exception as e:
        print("An error occurred in analyze_hairpin:", str(e))
        raise


def login_to_idt(driver):
    # Navigate to the login page
    driver.get("https://www.idtdna.com/site/account/login")
    #accept cookies
    try:
        cookies_accept_button = WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        )
        cookies_accept_button.click()
    except TimeoutException:
        print("No cookies acceptance button found or it timed out.")

    # Wait for the username input field to be present
    username_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "UserName"))  # Using the ID for the username field
    )
    
    # Wait for the password input field to be present
    password_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "Password"))  # Using the ID for the password field
    )

    # Fill in the username and password
    username_input.send_keys('NoahNicol')
    password_input.send_keys('nnicol101292')

    # Locate and check the "Remember Me" checkbox if not already checked
    remember_me_checkbox = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "RememberMe"))
    )
    if not remember_me_checkbox.is_selected():  # Check if the checkbox is not already checked
        remember_me_checkbox.click()

    # Wait for the login button to be clickable
    login_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "login-button"))
    )

    # Click the login button
    login_button.click()
    try:
        # Optionally, wait for a specific element that indicates the home page is loaded (when the user is logged in)
        WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "user-logged-in-btn"))
        )
    except:
        # A pause to ensure any redirects or loads complete if you're not waiting for a specific element after login
        time.sleep(8)
    print("Login sequence completed")


def update_excel_with_delta_g(file_path, sheet_name, dimer_results, hairpin_results, labels, params):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Populate labels 
    start_col = 11  # Column J
    for idx, label in enumerate(labels, start=start_col):
        sheet.cell(row=1, column=idx).value = label  # Populate top labels
        sheet.cell(row=idx, column=10).value = label  # Populate side labels
        sheet.cell(row=idx, column=9).value = hairpin_results[label]  # Populate hairpin results in column 9

    # Populate delta G values in the lower half of the matrix
    label_index = {label: idx + start_col for idx, label in enumerate(labels)}
    for (label1, label2), delta_g in dimer_results.items():
        row = label_index[label1]
        col = label_index[label2]
        
        # Fill only the lower half
        if row <= col:  # Ensure we're updating the lower triangle
            row, col = col, row
        sheet.cell(row=row, column=col).value = delta_g
    
    # Determine the starting row for parameter specifications to avoid overlap
    param_row = max(label_index.values()) + 2  # Start below the last row of the matrix

    # Adding parameter specifications
    default_text = 'default'
    sheet.cell(row=param_row, column=7).value = "Parameter Concentrations:"
    sheet.cell(row=param_row, column=9).value = f"Oligo: {params.get('oligo', default_text)}"
    sheet.cell(row=param_row + 1, column=9).value = f"Na: {params.get('na', default_text)}"
    sheet.cell(row=param_row + 2, column=9).value = f"Mg: {params.get('mg', default_text)}"
    sheet.cell(row=param_row + 3, column=9).value = f"dNTP: {params.get('dntp', default_text)}"

    workbook.save(file_path)
    workbook.close()

def main():
    file_path = '/Users/noahnicol/Desktop/Desktop/Job_Stuff/Proteovista/HSV_BlockAssembly1.xlsx'
    sheet_name = 'Block8'
    
    driver = get_driver()  # Setup WebDriver
    sequences, labels = read_sequences_from_excel(file_path, sheet_name)
    
    params = {'oligo': 0.25, 'na': 25, 'mg': 10, 'dntp': 0}
    dimer_results = {}
    hairpin_results = {}
    
    for i, seq1 in enumerate(sequences):
        for j, seq2 in enumerate(sequences):
            if i <= j:  # Only process each pair once
                delta_g = analyze_dimer(driver, seq1, seq2, **params)
                dimer_results[(labels[i], labels[j])] = delta_g
    
    for seq, label in zip(sequences, labels):
        hairpin_delta_g = analyze_hairpin(driver, seq, **params)
        hairpin_results[label] = hairpin_delta_g
    
    update_excel_with_delta_g(file_path, sheet_name, dimer_results, hairpin_results, labels, params)
    driver.quit()  # Close the browser

if __name__ == "__main__":
    main()
